"""
Export hasil screening SLR ke Excel + grafik visualisasi.
Menggabungkan metadata artikel (RIS) + keputusan AI (screening_log.json).

Usage:
  python export_excel.py
  python export_excel.py -i data/Ready\ \(1\).ris -l output/screening_log.json
  python export_excel.py -o hasil.xlsx
"""

import json
import io
import argparse
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path

def parse_ris(filepath):
    articles = []
    current = {}
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\n')
            if line.startswith('ER  -'):
                if current:
                    articles.append(current)
                    current = {}
            elif '  - ' in line:
                tag = line[:2].strip()
                value = line[6:]
                if tag == 'TY':
                    current = {'TY': value}
                elif tag in current:
                    if isinstance(current[tag], list):
                        current[tag].append(value)
                    else:
                        current[tag] = [current[tag], value]
                else:
                    current[tag] = value
    return articles

def flatten(val):
    if isinstance(val, list):
        return '; '.join(val)
    return val or ''

def fig_to_image(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    plt.close(fig)
    return buf

def make_charts(df):
    charts = {}
    colors_main = ['#2ecc71', '#e74c3c', '#95a5a6']
    colors_conf = ['#3498db', '#f39c12', '#e74c3c']

    # 1. Pie chart: Include vs Exclude
    fig, ax = plt.subplots(figsize=(6, 5))
    counts = df['Decision'].value_counts()
    labels = counts.index.tolist()
    sizes  = counts.values.tolist()
    color_map = {'INCLUDE': '#2ecc71', 'EXCLUDE': '#e74c3c', 'UNDECIDED': '#95a5a6'}
    clrs = [color_map.get(l, '#95a5a6') for l in labels]
    wedges, texts, autotexts = ax.pie(
        sizes, labels=labels, colors=clrs,
        autopct='%1.1f%%', startangle=90,
        wedgeprops={'edgecolor': 'white', 'linewidth': 2}
    )
    for at in autotexts:
        at.set_fontsize(11)
        at.set_fontweight('bold')
    ax.set_title('Hasil Screening\nInclude vs Exclude', fontsize=13, fontweight='bold', pad=15)
    charts['pie'] = fig_to_image(fig)

    # 2. Bar chart: Confidence level (include only)
    df_inc = df[df['Decision'] == 'INCLUDE']
    fig, ax = plt.subplots(figsize=(6, 5))
    conf_counts = df_inc['Confidence'].value_counts().reindex(['high', 'medium', 'low'], fill_value=0)
    bars = ax.bar(conf_counts.index, conf_counts.values, color=colors_conf, edgecolor='white', linewidth=1.5, width=0.5)
    for bar in bars:
        h = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, h + 0.3, str(int(h)), ha='center', fontweight='bold', fontsize=11)
    ax.set_title('Tingkat Kepercayaan AI\n(Artikel Include)', fontsize=13, fontweight='bold')
    ax.set_xlabel('Confidence Level', fontsize=11)
    ax.set_ylabel('Jumlah Artikel', fontsize=11)
    ax.set_ylim(0, conf_counts.max() + 5)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    charts['confidence'] = fig_to_image(fig)

    # 3. Bar chart: perbandingan jumlah include vs exclude
    fig, ax = plt.subplots(figsize=(6, 5))
    categories = ['Include', 'Exclude']
    values = [len(df[df['Decision']=='INCLUDE']), len(df[df['Decision']=='EXCLUDE'])]
    bars = ax.bar(categories, values, color=['#2ecc71', '#e74c3c'], edgecolor='white', linewidth=1.5, width=0.4)
    for bar in bars:
        h = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, h + 10, f'{int(h):,}', ha='center', fontweight='bold', fontsize=11)
    ax.set_title('Distribusi Hasil Screening\n(Total Artikel)', fontsize=13, fontweight='bold')
    ax.set_ylabel('Jumlah Artikel', fontsize=11)
    ax.set_ylim(0, max(values) * 1.1)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    charts['bar'] = fig_to_image(fig)

    # 4. Confidence breakdown: include + exclude side by side
    fig, ax = plt.subplots(figsize=(7, 5))
    conf_levels = ['high', 'medium', 'low']
    inc_conf = df[df['Decision']=='INCLUDE']['Confidence'].value_counts().reindex(conf_levels, fill_value=0)
    exc_conf = df[df['Decision']=='EXCLUDE']['Confidence'].value_counts().reindex(conf_levels, fill_value=0)
    x = range(len(conf_levels))
    width = 0.35
    b1 = ax.bar([i - width/2 for i in x], inc_conf.values, width, label='Include', color='#2ecc71', edgecolor='white')
    b2 = ax.bar([i + width/2 for i in x], exc_conf.values, width, label='Exclude', color='#e74c3c', edgecolor='white')
    ax.set_xticks(list(x))
    ax.set_xticklabels(conf_levels)
    ax.set_title('Confidence Level\nInclude vs Exclude', fontsize=13, fontweight='bold')
    ax.set_ylabel('Jumlah Artikel', fontsize=11)
    ax.legend()
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    charts['conf_compare'] = fig_to_image(fig)

    return charts

def parse_args():
    parser = argparse.ArgumentParser(description="Export hasil screening SLR ke Excel")
    parser.add_argument("-i", "--input",  default="sample.ris",                 help="Input RIS file (default: sample.ris)")
    parser.add_argument("-l", "--log",    default="output/screening_log.json",   help="Screening log JSON (default: output/screening_log.json)")
    parser.add_argument("-o", "--output", default="output/hasil_screening.xlsx", help="Output Excel file (default: output/hasil_screening.xlsx)")
    return parser.parse_args()


def main():
    args = parse_args()

    print("Membaca RIS...")
    articles = parse_ris(args.input)

    print("Membaca log screening...")
    with open(args.log, 'r') as f:
        log = json.load(f)

    rows = []
    for i, article in enumerate(articles):
        decision = log.get(str(i), {"decision": "undecided", "confidence": "-", "reason": "-"})
        rows.append({
            "No"         : i + 1,
            "Decision"   : decision.get("decision", "").upper(),
            "Confidence" : decision.get("confidence", ""),
            "Reason"     : decision.get("reason", ""),
            "Title"      : flatten(article.get("TI")),
            "Authors"    : flatten(article.get("AU")),
            "Year"       : flatten(article.get("PY") or article.get("Y1")),
            "Journal"    : flatten(article.get("JO") or article.get("JF") or article.get("T2")),
            "Abstract"   : flatten(article.get("AB")),
            "Keywords"   : flatten(article.get("KW")),
            "DOI"        : flatten(article.get("DO")),
            "URL"        : flatten(article.get("UR")),
        })

    df = pd.DataFrame(rows)

    print("Membuat grafik...")
    charts = make_charts(df)

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    print(f"Menulis ke {args.output}...")
    with pd.ExcelWriter(args.output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='All', index=False)
        df[df['Decision'] == 'INCLUDE'].to_excel(writer, sheet_name='Include', index=False)
        df[df['Decision'] == 'EXCLUDE'].to_excel(writer, sheet_name='Exclude', index=False)

        summary = pd.DataFrame([
            {"Kategori": "Total Artikel", "Jumlah": len(df)},
            {"Kategori": "Include",       "Jumlah": len(df[df['Decision']=='INCLUDE'])},
            {"Kategori": "Exclude",       "Jumlah": len(df[df['Decision']=='EXCLUDE'])},
            {"Kategori": "Undecided",     "Jumlah": len(df[df['Decision']=='UNDECIDED'])},
            {"Kategori": "Include Rate",  "Jumlah": f"{len(df[df['Decision']=='INCLUDE'])/len(df)*100:.1f}%"},
        ])
        summary.to_excel(writer, sheet_name='Summary', index=False)

        # Sheet Visualisasi
        writer.book.create_sheet('Visualisasi')
        ws = writer.sheets['Visualisasi']
        ws['A1'] = 'Visualisasi Hasil Screening SLR'
        ws['A1'].font = __import__('openpyxl').styles.Font(bold=True, size=14)

        positions = {
            'pie':          'A3',
            'bar':          'J3',
            'confidence':   'A30',
            'conf_compare': 'J30',
        }
        for key, cell in positions.items():
            img = XLImage(charts[key])
            ws.add_image(img, cell)

        # Auto-fit kolom di semua sheet kecuali Visualisasi
        for name, sheet in writer.sheets.items():
            if name == 'Visualisasi':
                continue
            for col in sheet.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    inc = len(df[df['Decision']=='INCLUDE'])
    exc = len(df[df['Decision']=='EXCLUDE'])
    print(f"\nSelesai! File: {args.output}")
    print(f"  - Sheet 'All'         : {len(df)} artikel")
    print(f"  - Sheet 'Include'     : {inc} artikel ({inc/len(df)*100:.1f}%)")
    print(f"  - Sheet 'Exclude'     : {exc} artikel")
    print(f"  - Sheet 'Summary'     : ringkasan statistik")
    print(f"  - Sheet 'Visualisasi' : 4 grafik")

if __name__ == "__main__":
    main()
